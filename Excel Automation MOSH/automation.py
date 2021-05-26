import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def percent_decrease(filename):

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row,3)
        corrected_price = (cell.value)*0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    naming = sheet.cell(1,4)
    naming.value = "Corrected Price"

    values = Reference(sheet,
              min_row = 2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save(filename)

    #cell = sheet['a1']
    #cell = sheet.cell(1,1)    #Row,Column
    #print("Label of the first cell : ",cell.value)
    #print("Maximum rows in the spreadsheet :",sheet.max_row) #max number of rows in the spreadsheet
